from openpyxl import load_workbook
import os
import re
import urllib2

# Load the workbook
wb = load_workbook(filename = 'machine-readable-url-puf.xlsx')
ws = wb.get_sheet_by_name('MR-PUF')

# Pull out all the URLs
urls = [[u[3].value, u[4].value] for u in ws.iter_rows(row_offset = 1) if u[4].value is not None and u[3].value is not None]

# Remove duplicate entries
existing = set([])
complete = set([])
for u in urls:
    # If it's not already in the set of existing urls...
    if u[1] not in existing:
        # Add it in and continue
        existing.add(u[1])
        complete.add((u[0], u[1]))

# Create the data file
if not os.path.exists('data'):
    os.makedirs('data')
else:
    print("data/ already exists, remove it before continuing.")
    quit()

# Create files to store the resulting JSON.
for u in complete:

    # There are some providers with the same name but in different
    # states with different urls. These providers will have a number
    # appended after their folder.
    append = 0

    # Create the file
    if not os.path.exists(os.path.join('data', u[0])):
        loc = os.path.join('data', u[0])
        os.makedirs(loc)
    else:
        append = 1
        while 1:
            try:
                loc = os.path.join('data', u[0] + str(append))
                os.makedirs(loc)
                break
            except OSError, e:
                append = append + 1

    # Download the JSON and place it in the file
    print("Downloading " + u[0] + ', ' + u[1])

    try:
        request = urllib2.Request(u[1])
        response = urllib2.urlopen(request, timeout = 16)
        json = response.read()
        file = open(os.path.join(loc, 'index.json'), 'w')
        file.write(json)
        file.close()
    except urllib2.HTTPError, e:
        print "HTTPError ({0}): {1}".format(e.errno, e.strerror)
        continue
    except urllib2.URLError, e:
        print "Connection timed out."
        continue
    except:
        print "Unknown error!"
        continue
